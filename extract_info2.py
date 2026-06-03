import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

print("\n--- CSE Class timetable 25-26 Even Sem S4- S8.xlsx ---")
try:
    wb = openpyxl.load_workbook('CSE Class timetable 25-26 Even Sem S4- S8.xlsx', data_only=True)
    for sheetname in wb.sheetnames:
        print(f"Sheet: {sheetname}")
        sheet = wb[sheetname]
        for row in sheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                print(row)
except Exception as e:
    print(f"Error reading xlsx: {e}")
